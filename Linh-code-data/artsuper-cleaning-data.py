import openpyxl
from openpyxl.utils import get_column_letter
import re # Import regex module

def process_ARTsuper_direct_modification(
    source_file_path,
    output_file_path="ARTsuper_modified.xlsx"
):
    """
    Directly modifies the ARTsuper.xlsx file:
    - Sets 'OptionName' column values to 'Balanced'.
    - Adds an 'Int/Ext' column (if not present) and sets its values based on 'Type' column content.
    - Adds a 'Fund Name' column (if not present) and sets its values to 'ARTsuper'.
    - Extracts "Internally Managed" or "Externally Managed" from 'Type' column and moves to 'Int/Ext',
      trimming the 'Type' column.
    - Deletes all rows from 5495 onwards (keeping row 5494 as the last data row).

    Args:
        source_file_path (str): The path to the original ARTsuper.xlsx file.
        output_file_path (str): The path to save the modified Excel file.
                                 Defaults to 'ARTsuper_modified.xlsx'.
    """
    print(f"\n[INFO] Starting direct modification of '{source_file_path}'...")
    
    try:
        wb = openpyxl.load_workbook(source_file_path)
        ws = wb.active # Get the active (first) sheet
        print(f"[DEBUG] Successfully loaded '{source_file_path}'. Processing sheet: '{ws.title}'")
    except FileNotFoundError:
        print(f"[ERROR] Source file '{source_file_path}' not found. Please check the path.")
        return
    except Exception as e:
        print(f"[ERROR] Error loading source workbook: {e}")
        return

    # --- Configuration ---
    HEADER_ROW_NUMBER = 1
    LAST_DATA_ROW_TO_KEEP = 5494 # We keep this row, delete from 5495 downwards

    OPTION_NAME_HEADER = "OptionName"
    OPTION_NAME_DEFAULT_VALUE = "Balanced"

    INT_EXT_HEADER = "Int/Ext"
    INT_EXT_DEFAULT_VALUE = "Externally Managed" # Fallback/default if not found in 'Type'
    INTERNALLY_MANAGED_VALUE = "Internally Managed" # Specific value to look for

    FUND_NAME_HEADER = "Fund Name"
    FUND_NAME_DEFAULT_VALUE = "ARTsuper"

    TYPE_HEADER = "Type" # Confirmed: Exact header for the 'Type' column

    # --- 1. Read existing headers and find/add column indices ---
    # First pass to get current headers
    headers = [cell.value for cell in ws[HEADER_ROW_NUMBER]]
    print(f"[DEBUG] Original Headers (Row {HEADER_ROW_NUMBER}): {headers}")

    # Initialize column indices (0-indexed for list operations)
    option_name_col_idx = -1
    int_ext_col_idx = -1
    fund_name_col_idx = -1
    type_col_idx = -1 

    # Find existing column indices (these might shift if new columns are inserted left of them)
    try:
        option_name_col_idx = headers.index(OPTION_NAME_HEADER)
    except ValueError:
        print(f"[ERROR] Header '{OPTION_NAME_HEADER}' not found in the first row. Please ensure it's exact.")
        wb.close()
        return
    
    try:
        type_col_idx = headers.index(TYPE_HEADER)
    except ValueError:
        print(f"[ERROR] Header '{TYPE_HEADER}' not found in the first row. Please ensure it's exact.")
        wb.close()
        return

    # --- Add 'Int/Ext' and 'Fund Name' columns if they don't exist ---
    # It's safest to add new columns to the end to avoid shifting issues with existing column indices
    # until all insertions are done and headers are re-read for final indices.
    
    # Check and add 'Int/Ext'
    if INT_EXT_HEADER not in headers:
        new_col_pos = ws.max_column + 1 # 1-based index: add after the last existing column
        ws.insert_cols(new_col_pos)
        ws.cell(row=HEADER_ROW_NUMBER, column=new_col_pos, value=INT_EXT_HEADER)
        print(f"[INFO] Added new column '{INT_EXT_HEADER}' at column {get_column_letter(new_col_pos)}")
        headers = [cell.value for cell in ws[HEADER_ROW_NUMBER]] # Re-read headers after insertion
    else:
        print(f"[INFO] Column '{INT_EXT_HEADER}' already exists.")

    # Check and add 'Fund Name'
    if FUND_NAME_HEADER not in headers:
        new_col_pos = ws.max_column + 1 # 1-based index: add after the last existing column (which might be the newly added Int/Ext)
        ws.insert_cols(new_col_pos)
        ws.cell(row=HEADER_ROW_NUMBER, column=new_col_pos, value=FUND_NAME_HEADER)
        print(f"[INFO] Added new column '{FUND_NAME_HEADER}' at column {get_column_letter(new_col_pos)}")
        headers = [cell.value for cell in ws[HEADER_ROW_NUMBER]] # Re-read headers after insertion
    else:
        print(f"[INFO] Column '{FUND_NAME_HEADER}' already exists.")
    
    # --- Get final 0-indexed positions of all necessary columns after all potential insertions ---
    option_name_col_idx = headers.index(OPTION_NAME_HEADER)
    int_ext_col_idx = headers.index(INT_EXT_HEADER)
    fund_name_col_idx = headers.index(FUND_NAME_HEADER)
    type_col_idx = headers.index(TYPE_HEADER) # Get updated index if columns to its left were inserted/shifted

    # Convert 0-indexed to 1-based for openpyxl ws.cell() operations
    option_name_ws_col = option_name_col_idx + 1 
    int_ext_ws_col = int_ext_col_idx + 1 
    fund_name_ws_col = fund_name_col_idx + 1 
    type_ws_col = type_col_idx + 1 

    print(f"[DEBUG] Final column indices for modifications (1-based for openpyxl):")
    print(f"  '{OPTION_NAME_HEADER}': {option_name_ws_col}")
    print(f"  '{INT_EXT_HEADER}': {int_ext_ws_col}")
    print(f"  '{FUND_NAME_HEADER}': {fund_name_ws_col}")
    print(f"  '{TYPE_HEADER}': {type_ws_col}")

    # --- 2. Iterate through data rows and update values ---
    rows_processed = 0
    start_data_row = HEADER_ROW_NUMBER + 1 # Start from the row after headers (Row 2)

    # Use ws.iter_rows for potentially better performance on large files if only reading specific cells,
    # but range(start, end) combined with ws.cell(row, col) is fine for direct cell writes.
    for row_num in range(start_data_row, LAST_DATA_ROW_TO_KEEP + 1):
        if row_num > ws.max_row: # Safety break if sheet ends prematurely
            print(f"[DEBUG] Reached end of sheet at row {row_num-1} before expected last data row ({LAST_DATA_ROW_TO_KEEP}).")
            break
            
        # Get original value from 'Type' column
        original_type_value = ws.cell(row=row_num, column=type_ws_col).value
        # Ensure it's a string and strip whitespace for robust processing
        processed_type_value = str(original_type_value).strip() if original_type_value is not None else ""
        
        int_ext_value_for_row = INT_EXT_DEFAULT_VALUE # Default fallback (Externally Managed)

        # Check for management type in 'Type' column (case-insensitive and whole word match using regex)
        # Using re.search with word boundaries (\b) and re.IGNORECASE for robust matching
        if re.search(r'\bInternally Managed\b', processed_type_value, re.IGNORECASE):
            int_ext_value_for_row = INTERNALLY_MANAGED_VALUE
            # Remove the phrase from 'Type' column, case-insensitively
            processed_type_value = re.sub(r'\bInternally Managed\b', '', processed_type_value, flags=re.IGNORECASE).strip()
        elif re.search(r'\bExternally Managed\b', processed_type_value, re.IGNORECASE):
            int_ext_value_for_row = INT_EXT_DEFAULT_VALUE # This is "Externally Managed"
            # Remove the phrase from 'Type' column, case-insensitively
            processed_type_value = re.sub(r'\bExternally Managed\b', '', processed_type_value, flags=re.IGNORECASE).strip()
        
        # Update cells in the current row
        ws.cell(row=row_num, column=option_name_ws_col, value=OPTION_NAME_DEFAULT_VALUE)
        ws.cell(row=row_num, column=int_ext_ws_col, value=int_ext_value_for_row)
        ws.cell(row=row_num, column=fund_name_ws_col, value=FUND_NAME_DEFAULT_VALUE)
        
        # Update the 'Type' column with the trimmed value
        # Set to None if empty after trimming, to clear cell contents
        ws.cell(row=row_num, column=type_ws_col, value=processed_type_value if processed_type_value else None) 

        rows_processed += 1
        if rows_processed % 1000 == 0:
            print(f"[DEBUG] Processed {rows_processed} data rows (up to Excel row {row_num}).")

    print(f"[INFO] Finished updating {rows_processed} data rows for '{OPTION_NAME_HEADER}', '{INT_EXT_HEADER}', '{FUND_NAME_HEADER}', and cleaning '{TYPE_HEADER}'.")

    # --- 3. Delete rows from 5495 onwards ---
    current_max_row_before_deletion = ws.max_row
    rows_to_delete_start = LAST_DATA_ROW_TO_KEEP + 1

    if rows_to_delete_start <= current_max_row_before_deletion:
        num_rows_to_delete = current_max_row_before_deletion - rows_to_delete_start + 1
        ws.delete_rows(rows_to_delete_start, num_rows_to_delete)
        print(f"[INFO] Deleted {num_rows_to_delete} rows starting from row {rows_to_delete_start}.")
    else:
        print(f"[INFO] No rows to delete after row {LAST_DATA_ROW_TO_KEEP}. Max row is {current_max_row_before_deletion}.")


    # --- 4. Save the modified workbook ---
    try:
        wb.save(output_file_path)
        print(f"[INFO] Successfully saved modified file to '{output_file_path}'.")
    except Exception as e:
        print(f"[ERROR] Error saving the modified Excel file: {e}")
    finally:
        if 'wb' in locals() and wb is not None:
            wb.close()

# =========================================================================
# Main execution block
# =========================================================================
if __name__ == "__main__":
    source_excel_file = r"D:\LinhDao\Programming\SUPERFUNdProject\ARTsuper.xlsx"
    output_excel_file = r"D:\LinhDao\Programming\SUPERFUNdProject\ARTsuper_modified.xlsx" # Saving to a new file for safety

    process_ARTsuper_direct_modification(
        source_file_path=source_excel_file,
        output_file_path=output_excel_file
    )