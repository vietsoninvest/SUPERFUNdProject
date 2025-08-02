import openpyxl
from openpyxl.utils import get_column_letter

def perform_column_operations(
    source_file_path,
    output_file_path
):
    """
    Performs column data transfer, renaming, and deletion on an Excel file.

    Args:
        source_file_path (str): Path to the input Excel file.
        output_file_path (str): Path to save the modified Excel file.
    """
    print(f"\n[INFO] Starting column operations on '{source_file_path}'...")
    
    try:
        wb = openpyxl.load_workbook(source_file_path)
        ws = wb.active # Get the active (first) sheet
        print(f"[DEBUG] Successfully loaded '{source_file_path}'. Processing sheet: '{ws.title}'")
    except FileNotFoundError:
        print(f"[ERROR] Source file '{source_file_path}' not found. Please check the path.")
        return
    except Exception as e:
        print(f"[ERROR] Error loading source workbook: {e}")
        return

    HEADER_ROW_NUMBER = 1 # Assuming header is always in the first row
    # Assuming LAST_DATA_ROW_TO_KEEP is consistent with previous script if needed for iteration bounds.
    # For now, we'll iterate up to ws.max_row, which should already be truncated to 5494.
    
    # --- Define Renaming Map ---
    # Keys are current header names, values are desired new header names
    rename_map = {
        "AsAtDate": "Effective Date",
        "OptionName": "Option Name",
        "Type": "Asset Class Name",
        "Name": "Name/Kind of Investment Item",
        "SecurityIdentifier": "Stock ID",
        "UnitsHeld": "Units Held",
        "Ownership": "% Ownership",
        "Value": "Value (AUD)",         # This 'Value' column will receive data from 'TotalValue'
        "Weighting": "Weighting",   # This 'Weighting' column will receive data from 'TotalWeighting'
    }

    # --- Define Columns to Delete ---
    columns_to_delete_names = [
        "ActualExposure",
        "EffectOfExposure",
        "TotalValue",       # Now data is transferred, this can be deleted
        "TotalWeighting",   # Now data is transferred, this can be deleted
        "TotalActualExposure",
    ]

    # --- 1. Identify current headers and find all necessary column indices ---
    current_headers = [cell.value for cell in ws[HEADER_ROW_NUMBER]]
    print(f"[DEBUG] Headers before any operations: {current_headers}")

    # Map header names to their 1-based column indices
    header_to_col_idx = {header: i + 1 for i, header in enumerate(current_headers)}

    # Get indices for data transfer
    total_value_col_idx = header_to_col_idx.get("TotalValue")
    value_col_idx = header_to_col_idx.get("Value")
    total_weighting_col_idx = header_to_col_idx.get("TotalWeighting")
    weighting_col_idx = header_to_col_idx.get("Weighting")

    # --- 2. Transfer Data from "TotalValue" to "Value" and "TotalWeighting" to "Weighting" ---
    data_transfer_count = 0
    start_data_row = HEADER_ROW_NUMBER + 1

    if total_value_col_idx and value_col_idx:
        print(f"[INFO] Transferring data from 'TotalValue' to 'Value' (cols {get_column_letter(total_value_col_idx)} to {get_column_letter(value_col_idx)}).")
        for row_num in range(start_data_row, ws.max_row + 1):
            source_value = ws.cell(row=row_num, column=total_value_col_idx).value
            ws.cell(row=row_num, column=value_col_idx, value=source_value)
            data_transfer_count += 1
    else:
        print("[WARNING] Could not find both 'TotalValue' and 'Value' columns for transfer. Skipping TotalValue to Value data transfer.")

    if total_weighting_col_idx and weighting_col_idx:
        print(f"[INFO] Transferring data from 'TotalWeighting' to 'Weighting' (cols {get_column_letter(total_weighting_col_idx)} to {get_column_letter(weighting_col_idx)}).")
        for row_num in range(start_data_row, ws.max_row + 1):
            source_value = ws.cell(row=row_num, column=total_weighting_col_idx).value
            ws.cell(row=row_num, column=weighting_col_idx, value=source_value)
            data_transfer_count += 1
    else:
        print("[WARNING] Could not find both 'TotalWeighting' and 'Weighting' columns for transfer. Skipping TotalWeighting to Weighting data transfer.")
    
    if data_transfer_count > 0:
        print(f"[INFO] Completed data transfer for {data_transfer_count // 2} rows (assuming 2 columns transferred per row).")

    # --- 3. Perform Column Renaming ---
    # Re-read headers after potential cell value changes, though header names themselves haven't changed yet.
    # We do this here to get the most up-to-date header mapping for renaming.
    current_headers_for_rename = [cell.value for cell in ws[HEADER_ROW_NUMBER]]
    
    for col_idx, header_value in enumerate(current_headers_for_rename):
        if header_value in rename_map:
            new_name = rename_map[header_value]
            ws.cell(row=HEADER_ROW_NUMBER, column=col_idx + 1, value=new_name)
            print(f"[DEBUG] Renamed '{header_value}' to '{new_name}' (Column {get_column_letter(col_idx + 1)})")
    
    # --- 4. Identify and Delete Columns ---
    # Get the updated headers *after* renaming to ensure we're looking for correct names
    # for deletion. Example: 'TotalValue' is still 'TotalValue' at this point for deletion.
    headers_after_rename = [cell.value for cell in ws[HEADER_ROW_NUMBER]]
    
    columns_to_delete_indices = [] # Store 1-based indices for deletion

    for col_name_to_delete in columns_to_delete_names:
        try:
            # Find the index of the column to delete in the *current* header list
            col_index = headers_after_rename.index(col_name_to_delete) + 1
            columns_to_delete_indices.append(col_index)
            print(f"[DEBUG] Identified column for deletion: '{col_name_to_delete}' at column {get_column_letter(col_index)}")
        except ValueError:
            print(f"[WARNING] Column '{col_name_to_delete}' not found for deletion. Skipping.")
    
    # Sort indices in descending order to avoid issues when deleting (delete higher index first)
    columns_to_delete_indices.sort(reverse=True)
    print(f"[DEBUG] Columns to delete (1-based indices, descending): {columns_to_delete_indices}")

    for col_idx_to_delete in columns_to_delete_indices:
        ws.delete_cols(col_idx_to_delete)
        print(f"[INFO] Deleted column {get_column_letter(col_idx_to_delete)}.")

    # --- Save the modified workbook ---
    try:
        wb.save(output_file_path)
        print(f"[INFO] Successfully saved modified file to '{output_file_path}'.")
    except Exception as e:
        print(f"[ERROR] Error saving the modified Excel file: {e}")
    finally:
        if 'wb' in locals() and wb is not None:
            wb.close()

# =========================================================================
# Main execution block
# =========================================================================
if __name__ == "__main__":
    # Input file will be the output from the previous script
    input_excel_file = r"D:\LinhDao\Programming\SUPERFUNdProject\ARTsuper_modified.xlsx"
    # Output for this step, ready for the finalization script
    output_excel_file = r"D:\LinhDao\Programming\SUPERFUNdProject\ARTsuper_final.xlsx" 

    perform_column_operations(
        source_file_path=input_excel_file,
        output_file_path=output_excel_file
    )

    print("\n[IMPORTANT NOTE ON EXECUTION ORDER]:")
    print("To get the full transformation, you need to run the scripts in sequence:")
    print("1. Run the initial modification script (e.g., your `process_ARTsuper_direct_modification.py`). This produces 'ARTsuper_modified.xlsx'.")
    print("2. Run *this* modified script (`perform_column_operations.py`) to transfer data, rename, and delete columns. This takes 'ARTsuper_modified.xlsx' as input and produces 'ARTsuper_final.xlsx'.")
    print("3. Then, run the last script (`finalize_artsuper_data.py`) which takes 'ARTsuper_final.xlsx' and produces 'Linh-artsuper-CleanedData.xlsx'.")