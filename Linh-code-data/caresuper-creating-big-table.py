import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import re
import csv
import os
import pandas as pd

# =========================================================================
# Helper Functions
# =========================================================================

def tokenize_string(s):
    """
    Cleans a string by removing non-alphanumeric characters (except spaces),
    converts it to lowercase, and splits it into a set of words (tokens).
    Handles NaN values by returning an empty set.
    """
    if pd.isna(s) or not isinstance(s, str):
        return set()
    cleaned_s = re.sub(r'[^a-z0-9\s]', '', str(s).lower()).strip()
    return set(cleaned_s.split())

def infer_country_from_text(text, country_keywords, australian_states):
    """
    Infers a country name from a given text string based on a list of keywords.
    Prioritizes a match with an Australian state/territory abbreviation.
    If no Australian state is found, it finds the longest country keyword match.
    
    Args:
        text (str): The string to search within (e.g., combined Address and Investment Name).
        country_keywords (dict): A dictionary where keys are lowercase country names
                                 and values are the original proper-cased country names.
        australian_states (list): A list of Australian state/territory abbreviations.
                                 
    Returns:
        str: The proper-cased country name if a match is found, otherwise None.
    """
    if pd.isna(text) or not isinstance(text, str):
        return None

    # First, check for Australian state abbreviations as a high-confidence signal
    for state in australian_states:
        if re.search(r'\b' + re.escape(state) + r'\b', text, re.IGNORECASE):
            return "Australia"
            
    text_tokens = tokenize_string(text)
    
    if not text_tokens:
        return None

    best_match = None
    max_common_tokens = 0

    # New logic: check for multi-word matches by comparing token sets
    for keyword_lower, original_name in country_keywords.items():
        keyword_tokens = set(keyword_lower.split())
        
        if keyword_tokens.issubset(text_tokens):
            num_tokens_matched = len(keyword_tokens)
            
            if num_tokens_matched > max_common_tokens:
                max_common_tokens = num_tokens_matched
                best_match = original_name
    
    return best_match

def load_country_codes_from_csv(file_path):
    """
    Loads country codes from a two-column CSV table (Code, Country Name)
    and returns a lookup dictionary.
    """
    country_codes = {}
    try:
        with open(file_path, mode='r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                name = row.get('Country')
                code = row.get('Code')
                if code and name:
                    country_codes[str(code).strip().upper()] = str(name).strip()
        print(f"[INFO] Successfully loaded {len(country_codes)} country codes from '{file_path}'.")
    except FileNotFoundError:
        print(f"[ERROR] Lookup file '{file_path}' not found. Listed Country column will not be updated.")
    except Exception as e:
        print(f"[ERROR] An error occurred while loading the lookup file: {e}")
    return country_codes

# =========================================================================
# Main Processing Function (Updated)
# =========================================================================
def process_source_excel_to_csv(
    source_file_path,
    destination_file="CleanedData.csv",
    country_code_lookup_file="D:\LinhDao\Programming\SUPERFUNdProject\InternationalCountryCodes.csv"
):
    """
    Processes a complicated source Excel file, extracts data, performs country inference,
    and populates a new CSV file.
    """
    print(f"\n[DEBUG] Starting to process source Excel file: '{source_file_path}'")
    
    # --- Load Source Workbook ---
    try:
        wb_source = openpyxl.load_workbook(source_file_path, data_only=True)
        source_sheet = wb_source.worksheets[0]
        if source_sheet.max_row < 2:
            print("[ERROR] Source sheet appears almost empty. Exiting.")
            wb_source.close()
            return
    except FileNotFoundError:
        print(f"[ERROR] Source file '{source_file_path}' not found. Please check the path.")
        return
    except Exception as e:
        print(f"[ERROR] Error loading source workbook: {e}")
        return

    # --- Load Lookup Data ---
    country_code_lookup = load_country_codes_from_csv(country_code_lookup_file)
    if not country_code_lookup:
        print("[WARNING] Country code lookup dictionary is empty. Country names will not be resolved.")

    # A list of Australian state/territory abbreviations for direct lookup
    australian_states = ["NSW", "NT", "QLD", "VIC", "SA", "TAS", "WA", "ACT"]

    # --- Step 1: First Pass to build a country reference list from Stock IDs ---
    print("[INFO] Performing first pass to build country reference list.")
    stock_id_countries = set()
    for row_cells in source_sheet.iter_rows():
        for cell in row_cells:
            cell_value = str(cell.value or '').strip().upper()
            if len(cell_value) >= 2:
                listed_country_code = cell_value[-2:]
                if listed_country_code in country_code_lookup:
                    stock_id_countries.add(country_code_lookup[listed_country_code])
    
    # --- Step 2: Build final country keyword dictionary for inference ---
    country_keywords_dict = {}
    for country in stock_id_countries:
        country_keywords_dict[country.lower()] = country
    # Add 'Australia' to the reference list as it's a known country
    country_keywords_dict['australia'] = 'Australia'
    
    print(f"[INFO] Built country inference dictionary with {len(country_keywords_dict)} unique countries.")

    # --- Configuration for Column Mapping and Data Extraction ---
    destination_columns_ordered = [
        "Effective Date", "Fund Name", "Option Name", "Asset Class Name",
        "Int/Ext", "Name/Kind of Investment Item", "Currency", "Stock ID",
        "Listed Country", "Units Held", "% Ownership", "Address",
        "Value (AUD)", "Weighting"
    ]
    HEADER_KEYWORDS = ["name", "value", "weighting"]
    TOTAL_KEYWORD = "total"
    def normalize_text(text):
        return str(text).strip().lower() if text is not None else ""
    source_to_dest_header_map = {
        'security identifier': 'Stock ID', 'stock id': 'Stock ID', 'id': 'Stock ID',
        'units held': 'Units Held', 'units': 'Units Held', 'value': 'Value (AUD)',
        'value (aud)': 'Value (AUD)', 'investment value': 'Value (AUD)',
        'investment value (aud)': 'Value (AUD)', 'weighting': 'Weighting',
        'weighting (%)': 'Weighting', 'weight': 'Weighting',
        'proportional weight': 'Weighting', 'currency': 'Currency',
        'currency code': 'Currency', '% ownership': '% Ownership',
        '% of property held': '% Ownership', 'listed country': 'Listed Country',
        'country': 'Listed Country', 'address': 'Address',
        'full address details': 'Address',
    }

    print(f"[INFO] Attempting to create or overwrite destination CSV file: '{destination_file}'")
    try:
        with open(destination_file, mode='w', newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            csv_writer.writerow(destination_columns_ordered)
            print(f"[INFO] Header written to '{destination_file}'. Starting data extraction.")

            rows_processed_count = 0
            tables_found_count = 0
            data_rows_transferred = 0
            current_asset_class_name = None
            current_int_ext = None
            current_effective_date = datetime.date(2024, 12, 31)
            current_state = 0
            current_table_column_indices = {}
            all_rows = list(source_sheet.iter_rows())

            for row_idx, row_cells in enumerate(all_rows):
                rows_processed_count += 1
                current_row_values = [normalize_text(cell.value) for cell in row_cells]

                if current_state == 0:
                    is_header_candidate = all(
                        any(kw in cell_val for cell_val in current_row_values) for kw in HEADER_KEYWORDS
                    )
                    if is_header_candidate:
                        int_ext_match = None
                        if row_idx > 0:
                            previous_row_combined_text = " ".join([normalize_text(cell.value) for cell in all_rows[row_idx-1]])
                            int_ext_match = re.search(r'\b(internally|externally)\s+managed\b', previous_row_combined_text, re.IGNORECASE)
                        if not int_ext_match and row_idx > 1:
                            second_previous_row_combined_text = " ".join([normalize_text(cell.value) for cell in all_rows[row_idx-2]])
                            int_ext_match = re.search(r'\b(internally|externally)\s+managed\b', second_previous_row_combined_text, re.IGNORECASE)
                        current_int_ext = int_ext_match.group(0).replace("managed", "").strip().title() if int_ext_match else "Externally"
                        asset_class_text = None
                        if row_idx > 0:
                            asset_class_text = " ".join([str(cell.value or '') for cell in all_rows[row_idx-1]]).strip()
                        if not asset_class_text and row_idx > 1:
                            asset_class_text = " ".join([str(cell.value or '') for cell in all_rows[row_idx-2]]).strip()
                        if asset_class_text:
                            cleaned_asset_class_text = re.sub(r'_x000d_|_x000a_', '', asset_class_text, flags=re.IGNORECASE)
                            cleaned_asset_class_text = re.sub(r'\b(total|portfolio|investment|class|items|details|breakdown|aud)\b', '', cleaned_asset_class_text, flags=re.IGNORECASE)
                            if int_ext_match:
                                cleaned_asset_class_text = re.sub(re.escape(int_ext_match.group(0)), '', cleaned_asset_class_text, flags=re.IGNORECASE)
                            cleaned_asset_class_text = re.sub(r'[:\-,.\[\]{}<>/&]+', ' ', cleaned_asset_class_text).strip()
                            cleaned_asset_class_text = re.sub(r'\s+', ' ', cleaned_asset_class_text).strip()
                            if cleaned_asset_class_text:
                                words = [p.strip() for p in cleaned_asset_class_text.split() if len(p.strip()) > 1 and not p.strip().isdigit()]
                                current_asset_class_name = " ".join(words[:2]).strip().title()
                        if not current_asset_class_name:
                            current_asset_class_name = "Unknown Asset Class"
                        print(f"  [INFO] Found Table at Row {row_idx + 1}. Context: Asset Class: '{current_asset_class_name}', Int/Ext: '{current_int_ext}'")
                        tables_found_count += 1
                        current_state = 2
                        current_table_column_indices = {}
                        found_source_col_indices = set()
                        for col_idx, cell_value in enumerate(current_row_values):
                            if cell_value.startswith("name") and col_idx not in found_source_col_indices:
                                current_table_column_indices['Name/Kind of Investment Item'] = col_idx
                                found_source_col_indices.add(col_idx)
                                break
                        for dest_col in destination_columns_ordered:
                            if dest_col in ["Effective Date", "Fund Name", "Option Name", "Asset Class Name", "Int/Ext", "Name/Kind of Investment Item"]:
                                continue
                            found_mapping_for_dest_col = False
                            for src_header_norm, mapped_dest_header in source_to_dest_header_map.items():
                                if mapped_dest_header == dest_col:
                                    for col_idx, cell_value in enumerate(current_row_values):
                                        if src_header_norm == cell_value and col_idx not in found_source_col_indices:
                                            current_table_column_indices[dest_col] = col_idx
                                            found_mapping_for_dest_col = True
                                            break
                                    if found_mapping_for_dest_col:
                                        break
                        if not current_table_column_indices:
                            print(f"    [WARNING] No mappable headers found in table at row {row_idx + 1}. Resetting state.")
                            current_state = 0
                
                elif current_state == 2:
                    is_total_row = any(cell_val == TOTAL_KEYWORD for cell_val in current_row_values)
                    data_to_append = [None] * len(destination_columns_ordered)
                    data_to_append[destination_columns_ordered.index("Effective Date")] = current_effective_date
                    data_to_append[destination_columns_ordered.index("Fund Name")] = "CareSuper"
                    data_to_append[destination_columns_ordered.index("Option Name")] = "Balanced"
                    data_to_append[destination_columns_ordered.index("Asset Class Name")] = current_asset_class_name
                    data_to_append[destination_columns_ordered.index("Int/Ext")] = 0 if current_int_ext == "Internally" else 1 if current_int_ext == "Externally" else None
                    row_has_meaningful_data = False
                    stock_id_value = None
                    address_value = None
                    name_value = None
                    
                    for dest_col_name, source_col_index in current_table_column_indices.items():
                        try:
                            cell_value = row_cells[source_col_index].value
                            processed_value = cell_value
                            if dest_col_name in ["Units Held", "Value (AUD)", "Weighting", "% Ownership"]:
                                if isinstance(cell_value, (int, float)):
                                    processed_value = cell_value
                                elif isinstance(cell_value, str):
                                    try:
                                        cleaned_str = cell_value.strip().replace(",", "").replace("$", "").replace("%", "")
                                        processed_value = float(cleaned_str) if cleaned_str else None
                                    except ValueError:
                                        processed_value = None
                                else:
                                    processed_value = None
                                if dest_col_name == "Weighting" and isinstance(processed_value, (int, float)):
                                    processed_value = processed_value * 1
                            elif dest_col_name == "Effective Date":
                                if isinstance(cell_value, (datetime.date, datetime.datetime)):
                                    processed_value = cell_value.date()
                                elif isinstance(cell_value, str):
                                    try:
                                        processed_value = datetime.datetime.strptime(cell_value, '%Y-%m-%d').date()
                                    except ValueError:
                                        try:
                                            processed_value = datetime.datetime.strptime(cell_value, '%m/%d/%Y').date()
                                        except ValueError:
                                            processed_value = None
                                else:
                                    processed_value = None
                            elif dest_col_name == "Name/Kind of Investment Item":
                                processed_value = "Sub Total" if isinstance(cell_value, str) and normalize_text(cell_value) == TOTAL_KEYWORD else str(cell_value).strip() if cell_value is not None else None
                                name_value = processed_value
                            elif dest_col_name == "Stock ID":
                                stock_id_value = str(cell_value).strip() if cell_value is not None else None
                                processed_value = stock_id_value
                            elif dest_col_name == "Address":
                                address_value = str(cell_value).strip() if cell_value is not None else None
                                processed_value = address_value
                            elif isinstance(processed_value, (str, int, float, datetime.date, datetime.datetime)):
                                processed_value = str(processed_value).strip() if processed_value is not None else None
                            else:
                                processed_value = None
                            dest_col_index = destination_columns_ordered.index(dest_col_name)
                            data_to_append[dest_col_index] = processed_value
                            if is_total_row:
                                row_has_meaningful_data = True
                            elif dest_col_name not in ["Effective Date", "Fund Name", "Option Name", "Asset Class Name", "Int/Ext"]:
                                if processed_value is not None and str(processed_value).strip() != "":
                                    row_has_meaningful_data = True
                        except Exception as e:
                            data_to_append[destination_columns_ordered.index(dest_col_name)] = None
                    
                    # New Logic: Country inference
                    listed_country_value = None
                    
                    # 1. Lookup Country from Stock ID (High Confidence)
                    if stock_id_value and len(stock_id_value) >= 2:
                        listed_country_code = stock_id_value[-2:].upper()
                        if listed_country_code in country_code_lookup:
                            listed_country_value = country_code_lookup[listed_country_code]
                        elif listed_country_code in australian_states:
                            listed_country_value = "Australia"
                        else:
                            # If no lookup is found, we might want to keep the code,
                            # but we'll let the inference logic handle it if needed.
                            listed_country_value = None

                    # 2. If no Listed Country from Stock ID, try inference from Address/Name
                    if not listed_country_value and address_value:
                        combined_text = f"{address_value} {name_value}"
                        inferred_country = infer_country_from_text(combined_text, country_keywords_dict, australian_states)
                        if inferred_country:
                            listed_country_value = inferred_country

                    # 3. Handle the case where country lookup/inference worked
                    if listed_country_value:
                        data_to_append[destination_columns_ordered.index("Listed Country")] = listed_country_value
                    
                    if row_has_meaningful_data:
                        csv_writer.writerow(data_to_append)
                        data_rows_transferred += 1
                    if is_total_row:
                        current_state = 0
                        current_asset_class_name = None
                        current_int_ext = None
                        current_table_column_indices = {}

            print(f"\n[INFO] Processing complete.")
            print(f"[INFO] Total rows processed in source sheet: {rows_processed_count}")
            print(f"[INFO] Total tables identified: {tables_found_count}")
            print(f"[INFO] Total data rows transferred to CSV: {data_rows_transferred}")

    except Exception as e:
        print(f"[ERROR] An error occurred while writing to the CSV file: {e}")
    finally:
        if 'wb_source' in locals() and wb_source is not None:
            wb_source.close()

# =========================================================================
# Main execution block
# =========================================================================
if __name__ == "__main__":
    cleaned_data_filename = "Linh-caresuper-CleanedData.csv"
    source_file_path = r"D:\LinhDao\Programming\SUPERFUNdProject\Linh-caresuper_splitted_tables.xlsx"
    country_code_lookup_file_path = r"D:\LinhDao\Programming\SUPERFUNdProject\InternationalCountryCodes.csv"

    print(f"[INFO] Attempting to process source file '{source_file_path}' and populate '{cleaned_data_filename}'...")
    process_source_excel_to_csv(
        source_file_path=source_file_path,
        destination_file=cleaned_data_filename,
        country_code_lookup_file=country_code_lookup_file_path
    )