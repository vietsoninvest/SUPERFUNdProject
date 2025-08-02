import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.utils import get_column_letter
import datetime
import re

# =========================================================================
# Function 1: create_excel_table_fixed
# =========================================================================

def create_excel_table_fixed(file_name="CleanedData.xlsx"):
    """
    Creates an Excel file named 'CleanedData.xlsx' with an empty table
    named 'CleanedData' and 14 specified columns.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CleanedDataSheet"

    column_headers = [
        "Effective Date", "Fund Name", "Option Name", "Asset Class Name",
        "Int/Ext", "Name/Kind of Investment Item", "Currency", "Stock Id",
        "Listed Country", "Units Held", "% Ownership", "Address",
        "Value(AUD)", "Weighting(%)"
    ]

    ws.append(column_headers)

    num_columns = len(column_headers)
    end_column_letter = get_column_letter(num_columns)
    table_range = f"A1:{end_column_letter}1" 

    tab = Table(displayName="CleanedData", ref=table_range)

    for i, header_name in enumerate(column_headers):
        tab.tableColumns.append(TableColumn(id=i+1, name=header_name))

    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    ws.add_table(tab)

    try:
        wb.save(file_name)
        print(f"Excel file '{file_name}' with table 'CleanedData' template created successfully.")
    except Exception as e:
        print(f"Error saving the Excel file template: {e}")

# =========================================================================
# Function 2: map_excel_data_openpyxl_only (Asset Class Name fix for "Sub Total" rows)
# =========================================================================

def map_excel_data_openpyxl_only(source_file_path, destination_file_path):
    """
    Maps data from a source Excel file to a destination Excel file
    using only openpyxl, handling specific column renames and "Sub Total" rows.
    Processes only the first sheet of the source file, with case-insensitive
    header matching.

    Args:
        source_file_path (str): Path to the Excel file containing all the information.
        destination_file_path (str): Path to the Excel file that acts as the template
                                     and will receive the mapped data.
    """
    try:
        # Load the source workbook with data_only=True
        print(f"DEBUG: Attempting to load source file: '{source_file_path}'")
        source_wb = openpyxl.load_workbook(source_file_path, data_only=True) 
        source_ws = source_wb.worksheets[0] 
        print(f"DEBUG: Successfully loaded source workbook. Processing sheet: '{source_ws.title}'.")

        # --- GLOBAL PRE-PROCESSING: Treat "-" as empty ---
        print("DEBUG: Starting global pre-processing: treating all cells containing only '-' as empty.")
        for row in source_ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == "-":
                    cell.value = None
                    # print(f"DEBUG: Changed cell {cell.coordinate} from '-' to None.") # Uncomment for verbose debugging of this step
        print("DEBUG: Global pre-processing complete.")
        # --- END GLOBAL PRE-PROCESSING ---


        # Load the destination workbook
        print(f"DEBUG: Attempting to load destination template: '{destination_file_path}'")
        dest_wb = openpyxl.load_workbook(destination_file_path)
        dest_ws = dest_wb.active
        print(f"DEBUG: Successfully loaded destination workbook. Active sheet: {dest_ws.title}")

        # --- Read Headers and Clean for Case-Insensitive Matching ---
        source_headers_raw = [cell.value for cell in source_ws[2]] # Assuming headers are in row 2
        print("\nDEBUG: Raw (repr) source headers from Excel (for identifying hidden chars/spaces - from row 2):")
        for i, header in enumerate(source_headers_raw):
            print(f"  Column {i+1}: {repr(header)}")

        source_headers_cleaned = [str(h).strip().upper() if h is not None else "" for h in source_headers_raw]
        print(f"DEBUG: Source Headers (cleaned and uppercased for internal matching): {source_headers_cleaned}")

        if not any(source_headers_cleaned):
            print("ERROR: Source headers are empty or contain only whitespace. Check row 2 of your source file.")
            return
        
        dest_headers = [cell.value for cell in dest_ws[1]]
        dest_headers_cleaned = [str(h).strip() if h is not None else "" for h in dest_headers] 

        print(f"DEBUG: Destination Headers (raw): {dest_headers}")
        print(f"DEBUG: Destination Headers (cleaned): {dest_headers_cleaned}")

        if not any(dest_headers_cleaned):
            print("ERROR: Destination headers are empty or contain only whitespace. Check row 1 of your destination file.")
            pass # Allow to continue for template generation


        # --- Column Mapping Setup ---
        column_map = {
            "ASSET CLASS": "Asset Class Name",
            "INTERNALLY MANAGED OR EXTERNALLY MANAGED": "Int/Ext",
            "CURRENCY": "Currency",
            "SECURITY IDENTIFIER": "Stock Id",
            "ADDRESS": "Address",
            "LISTED COUNTRY": "Listed Country", 
            "% OWNERSHIP / PROPERTY HELD": "% Ownership",
            "UNITS HELD": "Units Held",
            "VALUE(AUD)": "Value(AUD)",
            "WEIGHTING(%)": "Weighting(%)",
            # We are NOT adding the multiple "Name" columns here directly for 1:1 mapping,
            # as their specific handling is done by the targeted loop below.
        }
        
        dest_header_to_col_idx = {header: idx + 1 for idx, header in enumerate(dest_headers_cleaned)}
        source_col_indices = {header: idx + 1 for idx, header in enumerate(source_headers_cleaned)}

        print(f"DEBUG: Mapped Source Column Indices (using cleaned/uppercased names): {source_col_indices}")
        print(f"DEBUG: Mapped Destination Header to Column Index (using cleaned names): {dest_header_to_col_idx}")
        
        critical_source_col_check = "ASSET CLASS"
        if critical_source_col_check not in source_col_indices:
            print(f"ERROR: Column '{critical_source_col_check}' (from column_map) *still* not found in source file headers.")
            print(f"  This means '{critical_source_col_check}' does not exactly match any of the cleaned/uppercased source headers.")
            print(f"  List of available source headers (cleaned & uppercased): {source_headers_cleaned}")
            print(f"  Please compare this list VERY CAREFULLY with '{critical_source_col_check}'. Look for extra spaces (e.g., 'ASSET  CLASS'), typos, or non-printable characters.")
            return

        # --- Determine the last row to process (one before "Total Investment Item") ---
        last_data_row_idx = source_ws.max_row # Default to last row
        asset_class_col_idx_src_for_total_check = source_col_indices.get("ASSET CLASS")
        
        if asset_class_col_idx_src_for_total_check is None:
            print("WARNING: 'ASSET CLASS' column not found, cannot determine 'Total Investment Item' row. Processing all rows up to max_row.")
        else:
            found_total_investment_item = False
            for r_idx in range(3, source_ws.max_row + 1): # Start from data rows
                cell_value = str(source_ws.cell(row=r_idx, column=asset_class_col_idx_src_for_total_check).value).strip()
                if "total investment item" in cell_value.lower():
                    last_data_row_idx = r_idx - 1 # Stop one row before this
                    found_total_investment_item = True
                    print(f"DEBUG: Found 'Total Investment Item' at row {r_idx}. Last data row to process will be {last_data_row_idx}.")
                    break
            if not found_total_investment_item:
                print("DEBUG: 'Total Investment Item' row not found. Processing all rows until end of sheet.")


        rows_to_append = []

        # --- Iterate and Process Source Data Rows ---
        # Loop up to last_data_row_idx
        print(f"DEBUG: Source sheet max_row: {source_ws.max_row}. Starting data processing from row 3 up to {last_data_row_idx}.")
        
        if last_data_row_idx < 3:
            print("WARNING: No data rows to process based on 'Total Investment Item' or general sheet size. Output will be empty except for headers.")
            dest_wb.save(destination_file_path)
            return

        for row_idx in range(3, last_data_row_idx + 1): # Loop from row 3 up to the determined last data row
            source_row_values = [cell.value for cell in source_ws[row_idx]]
            print(f"\nDEBUG: --- Processing Source Row {row_idx} ---")
            print(f"DEBUG: Raw source row values for this row (after global '-' pre-processing): {source_row_values}")
            
            new_row_data = [None] * len(dest_headers)

            # --- Apply fixed values for the whole table ---
            if "Effective Date" in dest_header_to_col_idx:
                new_row_data[dest_header_to_col_idx["Effective Date"] - 1] = "31 Dec 2024"
            if "Fund Name" in dest_header_to_col_idx:
                new_row_data[dest_header_to_col_idx["Fund Name"] - 1] = "AustralianRetirementSuper"
            if "Option Name" in dest_header_to_col_idx:
                new_row_data[dest_header_to_col_idx["Option Name"] - 1] = "Balanced"

            # --- Extract "ASSET CLASS" (and related Int/Ext/Name logic) ---
            asset_class_col_idx_src = source_col_indices.get("ASSET CLASS")
            source_asset_class_val_raw = "" 
            if asset_class_col_idx_src is not None and (asset_class_col_idx_src - 1) < len(source_row_values):
                val = source_row_values[asset_class_col_idx_src - 1]
                source_asset_class_val_raw = str(val).strip() if val is not None else ""
            
            source_asset_class_val_lower = source_asset_class_val_raw.lower() 

            print(f"DEBUG: Extracted Asset Class Value (raw): '{source_asset_class_val_raw}'")
            print(f"DEBUG: Extracted Asset Class Value (lower): '{source_asset_class_val_lower}'")

            # --- Modified logic for "Sub Total" rows (which become "Total" rows) ---
            if "sub total" in source_asset_class_val_lower:
                print(f"DEBUG: Identified 'Sub Total' row at index {row_idx}. Applying NEW special logic for format 'SUB TOTAL [Asset Class Name] [Int/Ext]'.")
                
                # --- MODIFIED: Ensure Asset Class Name is always derived ---
                int_ext_val = "Externally Managed" # Default value for Int/Ext
                temp_asset_class_string = source_asset_class_val_lower # Use a mutable copy

                if "internally" in temp_asset_class_string:
                    int_ext_val = "Internally Managed"
                    temp_asset_class_string = temp_asset_class_string.replace("internally", "")
                elif "externally" in temp_asset_class_string:
                    # int_ext_val is already "Externally Managed" by default, no change needed.
                    temp_asset_class_string = temp_asset_class_string.replace("externally", "")
                
                # Always derive asset_class_name_val by removing "sub total" from the modified string
                temp_str_for_asset_class = temp_asset_class_string.replace("sub total", "").strip()
                asset_class_name_val = temp_str_for_asset_class.title() # Convert to title case


                if "Int/Ext" in dest_header_to_col_idx:
                    new_row_data[dest_header_to_col_idx["Int/Ext"] - 1] = int_ext_val
                    print(f"DEBUG: For Sub Total row {row_idx}, Int/Ext value set to '{int_ext_val}'.")

                if "Asset Class Name" in dest_header_to_col_idx:
                    new_row_data[dest_header_to_col_idx["Asset Class Name"] - 1] = asset_class_name_val
                    print(f"DEBUG: For Sub Total row {row_idx}, Asset Class Name set to '{asset_class_name_val}'.")

                if "Name/Kind of Investment Item" in dest_header_to_col_idx:
                    new_row_data[dest_header_to_col_idx["Name/Kind of Investment Item"] - 1] = "Total" 

            else: # Regular data row logic
                if "Asset Class Name" in dest_header_to_col_idx:
                    new_row_data[dest_header_to_col_idx["Asset Class Name"] - 1] = source_asset_class_val_raw.title()
                
                # --- Int/Ext Logic with Default for REGULAR rows ---
                if "Int/Ext" in dest_header_to_col_idx: # Ensure destination column exists
                    int_ext_col_idx_src = source_col_indices.get("INTERNALLY MANAGED OR EXTERNALLY MANAGED")
                    
                    assigned_int_ext = None # Initialize as None
                    
                    if int_ext_col_idx_src is not None and (int_ext_col_idx_src - 1) < len(source_row_values):
                        # Get value after global pre-processing (so "-" is already None)
                        int_ext_source_val = str(source_row_values[int_ext_col_idx_src - 1]).strip().lower() if source_row_values[int_ext_col_idx_src - 1] is not None else ""
                        
                        if "internally" in int_ext_source_val:
                            assigned_int_ext = "Internally Managed"
                        elif "externally" in int_ext_source_val:
                            assigned_int_ext = "Externally Managed"
                        # If the source value is empty/None or contains neither, assigned_int_ext remains None

                    # Apply default if it's still None for regular rows
                    if assigned_int_ext is None:
                        new_row_data[dest_header_to_col_idx["Int/Ext"] - 1] = "Externally Managed"
                        print(f"DEBUG: For regular row {row_idx}, Int/Ext value was not explicitly set, defaulted to 'Externally Managed'.")
                    else:
                        new_row_data[dest_header_to_col_idx["Int/Ext"] - 1] = assigned_int_ext
                        print(f"DEBUG: For regular row {row_idx}, Int/Ext value set to '{assigned_int_ext}'.")


                # --- Logic for Name/Kind of Investment Item (using pre-processed data) ---
                name_kind_investment_item_set = False
                name_source_headers_to_check = [ 
                    "NAME / KIND OF INVESTMENT ITEM",
                    "NAME OF INSTITUTION",
                    "NAME OF ISSUER / COUNTERPARTY",
                    "NAME OF FUND MANAGER"
                ]

                print(f"DEBUG: Attempting to map 'Name/Kind of Investment Item' for row {row_idx}.")
                print(f"DEBUG: Source column indices available for Name check (from source_col_indices): {source_col_indices}")
                print(f"DEBUG: Headers being checked for Name (hardcoded list): {name_source_headers_to_check}")


                if "Name/Kind of Investment Item" in dest_header_to_col_idx:
                    for src_header_candidate in name_source_headers_to_check:
                        # Check if the candidate header exists in our discovered source headers
                        if src_header_candidate in source_col_indices:
                            s_idx = source_col_indices[src_header_candidate] - 1
                            if s_idx < len(source_row_values):
                                current_value = source_row_values[s_idx] # Value is already pre-processed (None if it was "-")
                                # Debug print the value as repr to catch empty strings or hidden chars
                                print(f"DEBUG: For '{src_header_candidate}' (source col_idx: {s_idx + 1}), raw value: {repr(current_value)}")
                                
                                # A simple check for None or empty string is now sufficient
                                if current_value is not None and str(current_value).strip() != "":
                                    new_row_data[dest_header_to_col_idx["Name/Kind of Investment Item"] - 1] = current_value 
                                    name_kind_investment_item_set = True
                                    print(f"DEBUG: Mapped '{src_header_candidate}' (value: '{current_value}') to 'Name/Kind of Investment Item'. Moving to next row.")
                                    break # Crucial: Stop after finding the first non-empty value
                                else:
                                    print(f"DEBUG: Value for '{src_header_candidate}' was empty/None. Not using.")
                            else:
                                print(f"DEBUG: Source column index {s_idx+1} for '{src_header_candidate}' is out of bounds for current row {row_idx}'s values. Skipping.")
                        else:
                            print(f"DEBUG: Source header '{src_header_candidate}' not found in source_col_indices. This means it's not present in your source Excel's header row 2 after cleaning.")
                
                if not name_kind_investment_item_set and "Name/Kind of Investment Item" in dest_header_to_col_idx:
                    new_row_data[dest_header_to_col_idx["Name/Kind of Investment Item"] - 1] = None
                    print(f"DEBUG: No suitable 'Name' column found with a non-empty value for 'Name/Kind of Investment Item' for row {row_idx}. Setting to None.")


            # --- Map other columns based on column_map (excluding those already handled specifically) ---
            columns_to_skip_if_already_set = [
                "Effective Date", "Fund Name", "Option Name",
                "Asset Class Name", "Int/Ext", "Name/Kind of Investment Item" # Ensure this is always skipped if already set
            ]

            for src_col_name_for_map, dest_col_name in column_map.items():
                if src_col_name_for_map in source_col_indices and dest_col_name in dest_header_to_col_idx:
                    src_idx = source_col_indices[src_col_name_for_map] - 1
                    dest_idx = dest_header_to_col_idx[dest_col_name] - 1
                    
                    # Only map if the destination column is not one of our specially handled ones,
                    # OR if it IS one of them, but its value is still None.
                    # This prevents overwriting fixed values or values set by specific logic (like Name/Kind of Inv Item).
                    if dest_col_name not in columns_to_skip_if_already_set or new_row_data[dest_idx] is None:
                        if src_idx < len(source_row_values):
                            current_value_generic_map = source_row_values[src_idx] # Value is already pre-processed (None if it was "-")
                            if current_value_generic_map is not None and str(current_value_generic_map).strip() != "":
                                new_row_data[dest_idx] = current_value_generic_map
                                source_header_actual = source_headers_raw[src_idx] if src_idx < len(source_headers_raw) else "UNKNOWN_SOURCE_COLUMN"
                                print(f"DEBUG: Mapped (generic loop) '{source_header_actual}' (value: '{current_value_generic_map}') to '{dest_col_name}' in new_row_data.")
                            else:
                                print(f"DEBUG: Value for '{src_col_name_for_map}' (generic map) was empty/None after pre-processing. Not mapping.")
                        else:
                            print(f"DEBUG: Source column index {src_idx} for '{src_col_name_for_map}' out of bounds for row {row_idx}. Skipping mapping for this column.")
            
            # This check ensures we only append rows that have *any* non-empty data in them
            if any(val is not None and str(val).strip() != "" for val in new_row_data):
                rows_to_append.append(new_row_data)
                print(f"DEBUG: Successfully prepared row for appending: {new_row_data}")
            else:
                print(f"DEBUG: Row {row_idx} resulted in empty/None data, NOT appending: {new_row_data}")


        print(f"\nDEBUG: Final count of rows prepared for appending: {len(rows_to_append)}")
        if not rows_to_append:
            print("WARNING: After processing all source rows, 'rows_to_append' is empty. The output file will likely be empty except for headers.")
            dest_wb.save(destination_file_path)
            return

        # Clear existing data rows in destination sheet (keep header)
        print(f"DEBUG: Current max row in destination BEFORE clearing: {dest_ws.max_row}. Clearing rows from {dest_ws.max_row} down to 2.")
        for r_idx in range(dest_ws.max_row, 1, -1):
            dest_ws.delete_rows(r_idx)
        print(f"DEBUG: Max row in destination AFTER clearing: {dest_ws.max_row}. (Should be 1)")


        # Append all processed rows to the destination worksheet
        print(f"DEBUG: Appending {len(rows_to_append)} rows to destination worksheet.")
        for row_data in rows_to_append:
            dest_ws.append(row_data)
        print(f"DEBUG: Max row in destination AFTER appending: {dest_ws.max_row}.")


        # Update the table range to include all new data
        current_max_row = dest_ws.max_row
        num_columns = len(dest_headers) 
        end_column_letter = get_column_letter(num_columns)
        updated_table_range = f"A1:{end_column_letter}{current_max_row}"

        table_name = "CleanedData"
        if table_name in dest_ws.tables:
            excel_table = dest_ws.tables[table_name]
            excel_table.ref = updated_table_range
            print(f"DEBUG: Table '{table_name}' range updated to '{excel_table.ref}'.")
        else:
            print(f"WARNING: Table '{table_name}' not found. Re-adding it with range '{updated_table_range}'.")
            new_tab = Table(displayName="CleanedData", ref=updated_table_range)
            for i, header_name in enumerate(dest_headers):
                new_tab.tableColumns.append(TableColumn(id=i+1, name=header_name))
            style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            new_tab.tableStyleInfo = style
            dest_ws.add_table(new_tab)


        # Save the modified destination workbook
        dest_wb.save(destination_file_path)
        print(f"Successfully mapped data from '{source_file_path}' to '{destination_file_path}'.")

    except FileNotFoundError:
        print(f"ERROR: One or both of the specified Excel files were not found.")
        print(f"  Source path: {source_file_path}")
        print(f"  Destination path: {destination_file_path}")
    except Exception as e:
        print(f"ERROR: An unexpected error occurred during mapping: {e}")
        import traceback
        traceback.print_exc()


# =========================================================================
# Main execution block
# =========================================================================

if __name__ == "__main__":
    # IMPORTANT: Update these paths to match your file locations
    cleaned_data_filename = r"D:\LinhDao\Programming\SUPERFUNdProject\Linh-retirementtrust-CleanedData.xlsx"
    source_file_path = r"D:\LinhDao\Programming\SUPERFUNdProject\Linh-retirementtrust.xlsx"

    print(f"\n--- Starting script execution ---")
    print(f"[INFO] Ensuring output template '{cleaned_data_filename}' is ready...")
    create_excel_table_fixed(cleaned_data_filename)

    print(f"\n[INFO] Starting data mapping from '{source_file_path}' to '{cleaned_data_filename}'...")
    map_excel_data_openpyxl_only(
        source_file_path=source_file_path,
        destination_file_path=cleaned_data_filename
    )
    print(f"\n--- Script execution finished ---")